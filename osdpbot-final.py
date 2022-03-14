import telebot
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#функция, которая делает график обедов по графику
def obed(grafik):
	book = openpyxl.open(grafik)

	sheets = ['ИЗ', 'Заведение', 'Дожим']
	wb = Workbook()
	dest_filename = 'График обедов.xlsx'
	ws1 = wb.create_sheet(sh)
	ws1.title = sh

	for sh in sheets: 
		sheet = book[sh]

		for i in range(1,sheet.max_row+1):
			cellval = sheet[i][1].value + sheet[i][3].value / 2
			cell = 'E' + str(i)
			sheet[cell] = cellval

		dict1 = {}
		for i in range(1,sheet.max_row+1):
			dict1[sheet[i][0].value] = sheet[i][4].value
			
		sorted_values = sorted(dict1.values()) # Sort the values
		sorted_values_uniq = []
		for i in sorted_values:
			if i not in sorted_values_uniq:
				sorted_values_uniq.append(i)
			if i in sorted_values_uniq:
				continue

		sort_dict1 = {}
		for i in sorted_values_uniq:
			for key in dict1.keys():
				if dict1[key] == i:
					sort_dict1[key] = i

		print(sort_dict1)

		#создаем шапку листа
		if sh == 'ИЗ':
			time_iz = ['09:15 - 09:55', '10:00 - 10:40', '10:45 - 11:25', '11:30 - 12:10', '12:15 - 12:55', '13:00 - 13:40', '13:45 - 14:25', '14:30 - 15:10', '15:15 - 15:55']
			strok = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
		if sh == 'Дожим' or sh == 'Заведение':
			time_iz = ['10:45 - 11:25', '11:30 - 12:10', '12:15 - 12:55', '13:00 - 13:40', '13:45 - 14:25', '14:30 - 15:10', '15:15 - 15:55']
			strok = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

		for i in range(len(strok)):
			cell2 = strok[i] + '1'
			ws1[cell2] = time_iz[i]


		maxobed = round(len(sort_dict1.keys())/len(strok))+1

		print(maxobed)


		# заполняем график поровну
		h = 2
		d = 0
		for key in sort_dict1.keys():
			#if d > (len(strok) - 1):
				#break


			if d >= (len(strok) - 1):
				#h = maxobed + 1
				d = len(strok) - 1
				cell4 = strok[d] + str(h)
				ws1[cell4] = key
				h += 1

			if h <= maxobed and d < (len(strok) - 1):
				cell4 = strok[d] + str(h)
				ws1[cell4] = key
				h += 1
				if h > maxobed:
					h = 2
					d += 1


	sheet = wb.sheetnames
	print(sheet)
	de = wb["Sheet"]
	wb.remove(de)

	wb.save(filename = dest_filename)


# Создаем экземпляр бота
bot = telebot.TeleBot('5144266319:AAFHwNw_bncM8GrO6fHq3jmVjsJXO6U_G8A')
# Функция, обрабатывающая команду /start
@bot.message_handler(commands=["start"])
def start(m, res=False):
    bot.send_message(m.chat.id, 'Я на связи. Отправь мне график')
# Получение сообщений от юзера
@bot.message_handler(content_types=["document"])

#
@bot.message_handler(content_types=['document'])
def handle_docs_photo(message):
	try:
	    chat_id = message.chat.id

	    file_info = bot.get_file(message.document.file_id)
	    downloaded_file = bot.download_file(file_info.file_path)

	    src = message.document.file_name;
	    with open(src, 'wb') as new_file:
	        new_file.write(downloaded_file)

	    bot.reply_to(message, "Пожалуй, я сохраню это")
	    obed(src)
	    #bot.send_document(message.chat.id, data=open('График обедов.xlsx', 'rb'))
	    bot.send_document(message.chat.id, document=open('График обедов.xlsx', 'rb'))
	    os.remove(src)
	    os.remove('График обедов.xlsx')

	except Exception as e:
	    bot.reply_to(message, e)



if __name__ == '__main__':
    bot.polling(none_stop=True, interval=0)